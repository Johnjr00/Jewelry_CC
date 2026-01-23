from __future__ import annotations

import csv
import io
import sqlite3
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

from dataclasses import dataclass
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from functools import wraps
from pathlib import Path
from typing import Dict, Optional, Tuple

from flask import (
    Flask, g, redirect, render_template, request, url_for, flash,
    session, Response, jsonify,
    abort,
    send_file,
    make_response
)
from werkzeug.security import generate_password_hash, check_password_hash

APP_DIR = Path(__file__).resolve().parent
DB_PATH = APP_DIR / "inventory.db"
ACTIVITY_LOG_TEMPLATE = os.path.join(os.path.dirname(__file__), "MASTER ACTIVITY LOG.xlsx")
DAILY_COUNT_TEMPLATE = os.path.join(os.path.dirname(__file__), "Daily Count Sheet.xlsx")

NEW_RECEIPTS_CODE = "NEW-RECEIPTS"
NEW_RECEIPTS_NAME = "New Receipts (Virtual)"
RETURNS_CODE = "RETURNS"
RETURNS_NAME = "Returns (Virtual)"

STORE_TZ = ZoneInfo("America/Phoenix")

ROLE_ADMIN = "admin"
ROLE_STAFF = "staff"
DEFAULT_LOCATION_NAME = "Main Store"

ACTION_RECEIVE = "RECEIVE"
ACTION_MOVE = "MOVE"
ACTION_SOLD = "SOLD"
ACTION_MISSING = "MISSING"
ACTION_RETURN = "RETURN"
ACTION_CASE_CREATE = "CASE_CREATE"
ACTION_CASE_DELETE = "CASE_DELETE"
ACTION_CASE_EDIT = "CASE_EDIT"
ACTION_USER_CREATE = "USER_CREATE"
ACTION_USER_DISABLE = "USER_DISABLE"

ITEM_CATEGORIES = [
    {
        "name": "Earring",
        "count_key": "earrings",
        "plural": "Earrings",
        "short": "E",
        "receive_order": 1,
        "count_order": 3,
    },
    {
        "name": "Ring",
        "count_key": "rings",
        "plural": "Rings",
        "short": "R",
        "receive_order": 2,
        "count_order": 2,
    },
    {
        "name": "Necklace",
        "count_key": "necklaces",
        "plural": "Necklaces",
        "short": "N",
        "receive_order": 3,
        "count_order": 4,
    },
    {
        "name": "Bracelet",
        "count_key": "bracelets",
        "plural": "Bracelets",
        "short": "B",
        "receive_order": 4,
        "count_order": 1,
    },
    {
        "name": "Other",
        "count_key": "other",
        "plural": "Other",
        "short": "O",
        "receive_order": 5,
        "count_order": 5,
    },
]


def _sorted_item_categories(key: str):
    return sorted(ITEM_CATEGORIES, key=lambda c: c[key])


ITEM_TYPES_ORDER = [c["name"] for c in _sorted_item_categories("receive_order")]
ALLOWED_ITEM_TYPES = set(ITEM_TYPES_ORDER)
COUNT_CATEGORIES = _sorted_item_categories("count_order")
COUNT_FIELDS = [c["count_key"] for c in COUNT_CATEGORIES]
RESERVE_COUNT_FIELDS = [f"reserve_{c['count_key']}" for c in COUNT_CATEGORIES]
DIAMOND_TEST_OPTIONS = {"Y", "N", "NRT"}
RETURN_DIAMOND_OPTIONS = {"Y", "N", "N/A"}
LOCATION_CASE = "CASE"
LOCATION_RESERVE = "RESERVE"
INVENTORY_LOCATIONS = {LOCATION_CASE, LOCATION_RESERVE}

app = Flask(__name__)
app.secret_key = "change-this-in-production"


def _virtual_case_code(base_code: str, location_id: Optional[int], default_location_id: Optional[int]) -> str:
    if not location_id or location_id == default_location_id:
        return base_code
    return f"{base_code}-{location_id}"



# ---------------- DB helpers ----------------
def get_db() -> sqlite3.Connection:
    # Ensure DB + tables exist even under WSGI / Windows services
    if not DB_PATH.exists():
        init_db()

    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON;")

        # If the DB file exists but tables don't (or were wiped), rebuild schema.
        try:
            conn.execute("SELECT 1 FROM users LIMIT 1;")
        except sqlite3.OperationalError:
            conn.close()
            init_db()
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON;")

        # Lightweight migrations (safe no-ops if already applied)
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS locations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL
                );
                """
            )
            default_location = conn.execute(
                "SELECT id FROM locations ORDER BY id LIMIT 1"
            ).fetchone()
            if not default_location:
                conn.execute(
                    "INSERT INTO locations (name, is_active, created_at) VALUES (?, 1, ?)",
                    (DEFAULT_LOCATION_NAME, utc_now()),
                )
                default_location = conn.execute(
                    "SELECT id FROM locations ORDER BY id LIMIT 1"
                ).fetchone()
            default_location_id = default_location["id"]

            locations = conn.execute("SELECT id FROM locations").fetchall()
            for loc in locations:
                returns_code = _virtual_case_code(RETURNS_CODE, loc["id"], default_location_id)
                conn.execute(
                    """
                    INSERT OR IGNORE INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at)
                    VALUES (?, ?, ?, 1, 1, ?)
                    """,
                    (returns_code, loc["id"], RETURNS_NAME, utc_now()),
                )

            user_cols = [r["name"] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
            if "location_id" not in user_cols:
                conn.execute("ALTER TABLE users ADD COLUMN location_id INTEGER")
                conn.execute("UPDATE users SET location_id=? WHERE location_id IS NULL", (default_location_id,))

            case_info = conn.execute("PRAGMA table_info(cases)").fetchall()
            case_cols = [r["name"] for r in case_info]
            if "location_id" not in case_cols:
                conn.execute("ALTER TABLE cases ADD COLUMN location_id INTEGER")
                conn.execute("UPDATE cases SET location_id=? WHERE location_id IS NULL", (default_location_id,))

            count_cols = [r["name"] for r in conn.execute("PRAGMA table_info(case_counts)").fetchall()]
            if "location_id" not in count_cols:
                conn.execute("ALTER TABLE case_counts ADD COLUMN location_id INTEGER")
                conn.execute("UPDATE case_counts SET location_id=? WHERE location_id IS NULL", (default_location_id,))

            hist_cols = [r["name"] for r in conn.execute("PRAGMA table_info(history)").fetchall()]
            if "location_id" not in hist_cols:
                conn.execute("ALTER TABLE history ADD COLUMN location_id INTEGER")
                conn.execute("UPDATE history SET location_id=? WHERE location_id IS NULL", (default_location_id,))

            cols = [r["name"] for r in conn.execute("PRAGMA table_info(case_counts)").fetchall()]
            if "other" not in cols:
                conn.execute("ALTER TABLE case_counts ADD COLUMN other INTEGER NOT NULL DEFAULT 0 CHECK(other >= 0)")
            reserve_cols = {
                "reserve_bracelets": "INTEGER NOT NULL DEFAULT 0 CHECK(reserve_bracelets >= 0)",
                "reserve_rings": "INTEGER NOT NULL DEFAULT 0 CHECK(reserve_rings >= 0)",
                "reserve_earrings": "INTEGER NOT NULL DEFAULT 0 CHECK(reserve_earrings >= 0)",
                "reserve_necklaces": "INTEGER NOT NULL DEFAULT 0 CHECK(reserve_necklaces >= 0)",
                "reserve_other": "INTEGER NOT NULL DEFAULT 0 CHECK(reserve_other >= 0)",
            }
            for col, ctype in reserve_cols.items():
                if col not in cols:
                    conn.execute(f"ALTER TABLE case_counts ADD COLUMN {col} {ctype}")
            if any(col not in cols for col in reserve_cols):
                conn.execute(
                    """
                    UPDATE case_counts
                    SET reserve_bracelets=0,
                        reserve_rings=0,
                        reserve_earrings=0,
                        reserve_necklaces=0,
                        reserve_other=0
                    """
                )

            case_pk = {r["name"]: r["pk"] for r in case_info}
            if case_pk.get("location_id", 0) == 0:
                conn.execute("PRAGMA foreign_keys = OFF;")
                conn.execute("ALTER TABLE cases RENAME TO cases_old;")
                conn.execute(
                    """
                    CREATE TABLE cases (
                        case_code TEXT NOT NULL,
                        location_id INTEGER NOT NULL,
                        case_name TEXT NOT NULL,
                        is_virtual INTEGER NOT NULL DEFAULT 0,
                        is_active INTEGER NOT NULL DEFAULT 1,
                        created_at TEXT NOT NULL,
                        PRIMARY KEY(case_code, location_id),
                        FOREIGN KEY(location_id) REFERENCES locations(id)
                    );
                    """
                )
                conn.execute(
                    """
                    INSERT INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at)
                    SELECT case_code, location_id, case_name, is_virtual, is_active, created_at
                    FROM cases_old;
                    """
                )
                conn.execute("DROP TABLE cases_old;")
                conn.execute("PRAGMA foreign_keys = ON;")

            count_fks = conn.execute("PRAGMA foreign_key_list(case_counts)").fetchall()
            needs_case_counts_fk = not any(
                fk["table"] == "cases" and fk["from"] == "location_id" for fk in count_fks
            )
            if needs_case_counts_fk:
                conn.execute("PRAGMA foreign_keys = OFF;")
                conn.execute("ALTER TABLE case_counts RENAME TO case_counts_old;")
                conn.execute(
                    """
                    CREATE TABLE case_counts (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        ts_utc TEXT NOT NULL,
                        local_date TEXT NOT NULL,
                        case_code TEXT NOT NULL,
                        location_id INTEGER NOT NULL,
                        user_id INTEGER,
                        username TEXT,
                        bracelets INTEGER NOT NULL CHECK(bracelets >= 0),
                        rings INTEGER NOT NULL CHECK(rings >= 0),
                        earrings INTEGER NOT NULL CHECK(earrings >= 0),
                        necklaces INTEGER NOT NULL CHECK(necklaces >= 0),
                        other INTEGER NOT NULL DEFAULT 0 CHECK(other >= 0),
                        reserve_bracelets INTEGER NOT NULL DEFAULT 0 CHECK(reserve_bracelets >= 0),
                        reserve_rings INTEGER NOT NULL DEFAULT 0 CHECK(reserve_rings >= 0),
                        reserve_earrings INTEGER NOT NULL DEFAULT 0 CHECK(reserve_earrings >= 0),
                        reserve_necklaces INTEGER NOT NULL DEFAULT 0 CHECK(reserve_necklaces >= 0),
                        reserve_other INTEGER NOT NULL DEFAULT 0 CHECK(reserve_other >= 0),
                        total INTEGER NOT NULL CHECK(total >= 0),
                        notes TEXT,
                        FOREIGN KEY(case_code, location_id) REFERENCES cases(case_code, location_id),
                        FOREIGN KEY(location_id) REFERENCES locations(id),
                        FOREIGN KEY(user_id) REFERENCES users(id)
                    );
                    """
                )
                conn.execute(
                    """
                    INSERT INTO case_counts (
                        id, ts_utc, local_date, case_code, location_id, user_id, username,
                        bracelets, rings, earrings, necklaces, other,
                        reserve_bracelets, reserve_rings, reserve_earrings, reserve_necklaces, reserve_other,
                        total, notes
                    )
                    SELECT
                        id, ts_utc, local_date, case_code, location_id, user_id, username,
                        bracelets, rings, earrings, necklaces, other,
                        reserve_bracelets, reserve_rings, reserve_earrings, reserve_necklaces, reserve_other,
                        total, notes
                    FROM case_counts_old;
                    """
                )
                conn.execute("DROP TABLE case_counts_old;")
                conn.execute(
                    "CREATE INDEX IF NOT EXISTS idx_case_counts_date_case ON case_counts(local_date, case_code, location_id);"
                )
                conn.execute(
                    "CREATE INDEX IF NOT EXISTS idx_case_counts_case ON case_counts(case_code, location_id);"
                )
                conn.execute("PRAGMA foreign_keys = ON;")
        except sqlite3.OperationalError:
            pass

        try:
            inv_info = conn.execute("PRAGMA table_info(inventory)").fetchall()
            inv_cols = [r["name"] for r in inv_info]
            inv_pk = {r["name"]: r["pk"] for r in inv_info}
            needs_inventory_rebuild = (
                "location_id" not in inv_cols
                or inv_pk.get("location_id", 0) == 0
                or inv_pk.get("case_code", 0) == 0
                or inv_pk.get("upc", 0) == 0
                or inv_pk.get("location", 0) == 0
            )
            if needs_inventory_rebuild:
                conn.execute("PRAGMA foreign_keys = OFF;")
                conn.execute("ALTER TABLE inventory RENAME TO inventory_old;")
                conn.execute(
                    """
                    CREATE TABLE inventory (
                        case_code TEXT NOT NULL,
                        location_id INTEGER NOT NULL,
                        upc TEXT NOT NULL,
                        location TEXT NOT NULL DEFAULT 'CASE' CHECK(location IN ('CASE','RESERVE')),
                        qty INTEGER NOT NULL CHECK(qty >= 0),
                        PRIMARY KEY(case_code, location_id, upc, location),
                        FOREIGN KEY(case_code, location_id) REFERENCES cases(case_code, location_id),
                        FOREIGN KEY(upc) REFERENCES products(upc)
                    );
                    """
                )
                if "location" in inv_cols:
                    if "location_id" in inv_cols:
                        conn.execute(
                            """
                            INSERT INTO inventory (case_code, location_id, upc, location, qty)
                            SELECT case_code, location_id, upc, location, qty
                            FROM inventory_old;
                            """
                        )
                    else:
                        conn.execute(
                            """
                            INSERT INTO inventory (case_code, location_id, upc, location, qty)
                            SELECT inv.case_code,
                                   COALESCE(c.location_id, ?) AS location_id,
                                   inv.upc,
                                   inv.location,
                                   inv.qty
                            FROM inventory_old inv
                            LEFT JOIN cases c ON c.case_code = inv.case_code;
                            """,
                            (default_location_id,),
                        )
                else:
                    conn.execute(
                        """
                        INSERT INTO inventory (case_code, location_id, upc, location, qty)
                        SELECT inv.case_code,
                               COALESCE(c.location_id, ?) AS location_id,
                               inv.upc,
                               'CASE' AS location,
                               inv.qty
                        FROM inventory_old inv
                        LEFT JOIN cases c ON c.case_code = inv.case_code;
                        """,
                        (default_location_id,),
                    )
                conn.execute("DROP TABLE inventory_old;")
                conn.execute("PRAGMA foreign_keys = ON;")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_inv_case ON inventory(case_code, location_id);")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_inv_upc ON inventory(upc);")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_inv_case_location ON inventory(case_code, location_id, location);")
        except sqlite3.OperationalError:
            pass

        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def local_date_str() -> str:
    # Store-local date for daily counts
    return datetime.now(STORE_TZ).date().isoformat()

def now_local_dt() -> datetime:
    """Store-local datetime (timezone-aware)."""
    return datetime.now(STORE_TZ)


def store_now() -> datetime:
    """Alias for store-local now (timezone-aware)."""
    return datetime.now(STORE_TZ)




def _parse_iso_utc(value: str) -> Optional[datetime]:
    if not value:
        return None
    s = str(value).strip().replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(s)
    except Exception:
        try:
            dt = datetime.fromisoformat(s.split(".")[0])
        except Exception:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt

def fmt_local_ts(value: str) -> str:
    """Format ISO UTC timestamp into store-local MM/DD/YYYY HH:MM AM/PM."""
    dt = _parse_iso_utc(value)
    if not dt:
        return ""
    return dt.astimezone(STORE_TZ).strftime("%m/%d/%Y %I:%M %p")

def fmt_mmddyyyy(value: str) -> str:
    """Format YYYY-MM-DD (or ISO timestamp) into MM/DD/YYYY."""
    if not value:
        return ""
    s = str(value).strip()
    # If it's a full timestamp, use local_ts and take date part
    if "T" in s:
        dt = _parse_iso_utc(s)
        if not dt:
            return ""
        return dt.astimezone(STORE_TZ).strftime("%m/%d/%Y")
    try:
        d = datetime.fromisoformat(s).date()
        return d.strftime("%m/%d/%Y")
    except Exception:
        # try MM/DD/YYYY passthrough
        try:
            d = datetime.strptime(s, "%m/%d/%Y").date()
            return d.strftime("%m/%d/%Y")
        except Exception:
            return s

# Jinja filters
app.jinja_env.filters["local_ts"] = fmt_local_ts
app.jinja_env.filters["mmddyyyy"] = fmt_mmddyyyy

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON;")

    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS locations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS cases (
            case_code TEXT NOT NULL,
            location_id INTEGER NOT NULL,
            case_name TEXT NOT NULL,
            is_virtual INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            PRIMARY KEY(case_code, location_id),
            FOREIGN KEY(location_id) REFERENCES locations(id)
        );

        CREATE TABLE IF NOT EXISTS products (
            upc TEXT PRIMARY KEY,
            description TEXT
        );

        CREATE TABLE IF NOT EXISTS inventory (
            case_code TEXT NOT NULL,
            location_id INTEGER NOT NULL,
            upc TEXT NOT NULL,
            location TEXT NOT NULL DEFAULT 'CASE' CHECK(location IN ('CASE','RESERVE')),
            qty INTEGER NOT NULL CHECK(qty >= 0),
            PRIMARY KEY(case_code, location_id, upc, location),
            FOREIGN KEY(case_code, location_id) REFERENCES cases(case_code, location_id),
            FOREIGN KEY(upc) REFERENCES products(upc)
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin','staff')),
            is_active INTEGER NOT NULL DEFAULT 1,
            location_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(location_id) REFERENCES locations(id)
        );


CREATE TABLE IF NOT EXISTS case_counts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ts_utc TEXT NOT NULL,
    local_date TEXT NOT NULL,
    case_code TEXT NOT NULL,
    location_id INTEGER NOT NULL,
    user_id INTEGER,
    username TEXT,
    bracelets INTEGER NOT NULL CHECK(bracelets >= 0),
    rings INTEGER NOT NULL CHECK(rings >= 0),
    earrings INTEGER NOT NULL CHECK(earrings >= 0),
    necklaces INTEGER NOT NULL CHECK(necklaces >= 0),
    other INTEGER NOT NULL DEFAULT 0 CHECK(other >= 0),
    reserve_bracelets INTEGER NOT NULL DEFAULT 0 CHECK(reserve_bracelets >= 0),
    reserve_rings INTEGER NOT NULL DEFAULT 0 CHECK(reserve_rings >= 0),
    reserve_earrings INTEGER NOT NULL DEFAULT 0 CHECK(reserve_earrings >= 0),
    reserve_necklaces INTEGER NOT NULL DEFAULT 0 CHECK(reserve_necklaces >= 0),
    reserve_other INTEGER NOT NULL DEFAULT 0 CHECK(reserve_other >= 0),
    total INTEGER NOT NULL CHECK(total >= 0),
    notes TEXT,
    FOREIGN KEY(case_code, location_id) REFERENCES cases(case_code, location_id),
    FOREIGN KEY(location_id) REFERENCES locations(id),
    FOREIGN KEY(user_id) REFERENCES users(id)
);

CREATE INDEX IF NOT EXISTS idx_case_counts_date_case ON case_counts(local_date, case_code, location_id);
CREATE INDEX IF NOT EXISTS idx_case_counts_case ON case_counts(case_code, location_id);

        CREATE TABLE IF NOT EXISTS history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ts TEXT NOT NULL,
    user_id INTEGER,
    username TEXT,
    action TEXT NOT NULL,
    upc TEXT,
    qty INTEGER,
    from_case_code TEXT,
    to_case_code TEXT,
    notes TEXT,
    trans_reg TEXT,
    dept_no TEXT,
    brief_desc TEXT,
    ticket_price REAL,
    diamond_test TEXT,
    location_id INTEGER,
    FOREIGN KEY(user_id) REFERENCES users(id),
    FOREIGN KEY(location_id) REFERENCES locations(id)
);

        CREATE INDEX IF NOT EXISTS idx_inv_case ON inventory(case_code, location_id);
        CREATE INDEX IF NOT EXISTS idx_inv_upc ON inventory(upc);
        CREATE INDEX IF NOT EXISTS idx_inv_case_location ON inventory(case_code, location_id, location);
        CREATE INDEX IF NOT EXISTS idx_hist_upc ON history(upc);
        CREATE INDEX IF NOT EXISTS idx_hist_case_from ON history(from_case_code);
        CREATE INDEX IF NOT EXISTS idx_hist_case_to ON history(to_case_code);
        """
    )

    # --- Migration: add products.item_type if missing ---
    cols = [r["name"] for r in db.execute("PRAGMA table_info(products)").fetchall()]
    if "item_type" not in cols:
        db.execute("ALTER TABLE products ADD COLUMN item_type TEXT;")
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS locations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );
        """
    )
    default_location = db.execute(
        "SELECT id FROM locations ORDER BY id LIMIT 1"
    ).fetchone()
    if not default_location:
        db.execute(
            "INSERT INTO locations (name, is_active, created_at) VALUES (?, 1, ?)",
            (DEFAULT_LOCATION_NAME, utc_now()),
        )
        default_location = db.execute(
            "SELECT id FROM locations ORDER BY id LIMIT 1"
        ).fetchone()
    default_location_id = default_location["id"]
    # Ensure New Receipts exists
    db.execute(
        """
        INSERT OR IGNORE INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at)
        VALUES (?, ?, ?, 1, 1, ?)
        """,
        (NEW_RECEIPTS_CODE, default_location_id, NEW_RECEIPTS_NAME, utc_now()),
    )
    locations = db.execute("SELECT id FROM locations").fetchall()
    for loc in locations:
        returns_code = _virtual_case_code(RETURNS_CODE, loc["id"], default_location_id)
        db.execute(
            """
            INSERT OR IGNORE INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at)
            VALUES (?, ?, ?, 1, 1, ?)
            """,
            (returns_code, loc["id"], RETURNS_NAME, utc_now()),
        )

    # --- Lightweight migration for history SOLD fields (safe on existing DBs)
    try:
        cols = [r["name"] for r in db.execute("PRAGMA table_info(history)").fetchall()]
        wanted = {
            "trans_reg": "TEXT",
            "dept_no": "TEXT",
            "brief_desc": "TEXT",
            "ticket_price": "REAL",
            "diamond_test": "TEXT",
        }
        for col, ctype in wanted.items():
            if col not in cols:
                db.execute(f"ALTER TABLE history ADD COLUMN {col} {ctype}")
    except Exception:
        pass

    db.commit()
    db.close()

def is_setup_complete() -> bool:
    """Returns True once at least 1 user exists."""
    try:
        db = get_db()
        c = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
        return c > 0
    except Exception:
        # On a brand new DB, init_db should have created tables; if it didn't, treat as not complete.
        return False

# ---------------- Auth helpers ----------------
@dataclass
class CurrentUser:
    id: int
    username: str
    role: str
    location_id: Optional[int]


def current_user() -> Optional[CurrentUser]:
    uid = session.get("user_id")
    if not uid:
        return None
    db = get_db()
    row = db.execute(
        "SELECT id, username, role, location_id FROM users WHERE id=? AND is_active=1",
        (uid,),
    ).fetchone()
    if not row:
        session.pop("user_id", None)
        return None
    return CurrentUser(
        id=row["id"],
        username=row["username"],
        role=row["role"],
        location_id=row["location_id"],
    )


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        # Force initial setup on brand-new installs
        if not is_setup_complete() and request.endpoint != "setup":
            return redirect(url_for("setup"))
        u = current_user()
        if not u:
            return redirect(url_for("login", next=request.path))
        if u.role == ROLE_ADMIN:
            if current_location_id():
                valid_location = get_db().execute(
                    "SELECT id FROM locations WHERE id=? AND is_active=1",
                    (current_location_id(),),
                ).fetchone()
                if not valid_location:
                    session.pop("location_id", None)
            if not current_location_id():
                if request.endpoint not in (
                    "select_location",
                    "locations_admin",
                    "logout",
                    "login",
                    "setup",
                    "static",
                ):
                    return redirect(url_for("select_location"))
        else:
            if not u.location_id:
                session.pop("user_id", None)
                flash("No store location assigned. Contact an admin.", "danger")
                return redirect(url_for("login"))
            session["location_id"] = u.location_id
        return fn(*args, **kwargs)
    return wrapper


def role_required(role: str):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            u = current_user()
            if not u:
                return redirect(url_for("login", next=request.path))
            if u.role != role:
                flash("You don’t have permission for that.", "danger")
                return redirect(url_for("index"))
            return fn(*args, **kwargs)
        return wrapper
    return deco


def current_location_id() -> Optional[int]:
    u = current_user()
    if not u:
        return None
    if u.role == ROLE_ADMIN:
        return session.get("location_id")
    return u.location_id


def current_location() -> Optional[sqlite3.Row]:
    loc_id = session.get("location_id")
    if not loc_id:
        return None
    db = get_db()
    return db.execute(
        "SELECT id, name FROM locations WHERE id=? AND is_active=1",
        (loc_id,),
    ).fetchone()


def new_receipts_case_code(location_id: Optional[int]) -> str:
    if not location_id:
        return NEW_RECEIPTS_CODE
    db = get_db()
    row = db.execute("SELECT id FROM locations ORDER BY id LIMIT 1").fetchone()
    default_location_id = row["id"] if row else None
    return _virtual_case_code(NEW_RECEIPTS_CODE, location_id, default_location_id)


def returns_case_code(location_id: Optional[int]) -> str:
    if not location_id:
        return RETURNS_CODE
    db = get_db()
    row = db.execute("SELECT id FROM locations ORDER BY id LIMIT 1").fetchone()
    default_location_id = row["id"] if row else None
    return _virtual_case_code(RETURNS_CODE, location_id, default_location_id)


@app.context_processor
def inject_location():
    return {"current_location": current_location()}



# ---------------- Daily Counts ----------------
def _initials_from_username(username: str) -> str:
    if not username:
        return ""
    parts = re.split(r"\s+", username.strip())
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()
    u = re.sub(r"[^A-Za-z0-9]", "", username)
    return (u[:2] if len(u) >= 2 else u).upper()

def _item_code(item_type: str) -> str:
    m = {
        "ring": "R",
        "earring": "E",
        "earrings": "E",
        "necklace": "N",
        "bracelet": "B",
        "other": "O",
    }
    if not item_type:
        return "O"
    return m.get(item_type.strip().lower(), "O")

def _reason_code(action: str) -> str:
    a = (action or "").upper()
    # Matches the legend in the template
    if a == "RECEIVE":
        return "NRT"
    if a == "MOVE":
        return "M"
    if a == "SOLD":
        return "S"
    if a == "MISSING":
        return "D"  # closest match in template legend
    if a == "RETURN":
        return "R"
    return a[:3]

def _parse_iso_to_store(value: str):
    if not value:
        return None
    s = str(value).replace("Z", "")
    try:
        dt = datetime.fromisoformat(s)
    except Exception:
        try:
            dt = datetime.fromisoformat(s.split(".")[0])
        except Exception:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=STORE_TZ)
    return dt.astimezone(STORE_TZ)

def _local_date_str_from_ts(ts: str) -> str:
    dt = _parse_iso_to_store(ts)
    if not dt:
        return ""
    return dt.date().isoformat()

def _daily_activity_totals(case_code: str, local_date: str, location_id: int) -> tuple[int, int, int]:
    db = get_db()
    events = db.execute(
        """
        SELECT h.ts, h.action, h.qty, h.from_case_code, h.to_case_code
        FROM history h
        WHERE h.action IN ('RECEIVE','MOVE','SOLD','MISSING','RETURN')
          AND h.location_id = ?
          AND (h.from_case_code = ? OR h.to_case_code = ?)
        ORDER BY h.ts ASC, h.id ASC
        """,
        (location_id, case_code, case_code),
    ).fetchall()

    total_in = 0
    total_out = 0
    for e in events:
        ld = _local_date_str_from_ts(e["ts"])
        if ld != local_date:
            continue
        action = (e["action"] or "").upper()
        qty = int(e["qty"] or 0)
        qty_in = 0
        qty_out = 0
        if action == "RECEIVE":
            if e["to_case_code"] == case_code:
                qty_in = qty
            else:
                qty_out = qty
        elif action == "MOVE":
            if e["to_case_code"] == case_code:
                qty_in = qty
            elif e["from_case_code"] == case_code:
                qty_out = qty
        elif action == "RETURN":
            if e["to_case_code"] == case_code:
                qty_in = qty
            elif e["from_case_code"] == case_code:
                qty_out = qty
        else:
            if e["from_case_code"] == case_code:
                qty_out = qty
        total_in += qty_in
        total_out += qty_out

    return total_in, total_out, total_in - total_out

def build_daily_activity_workbook(case_code: str, local_date: str, location_id: int):
    """Return an openpyxl workbook for the given case and local_date (YYYY-MM-DD) using MASTER ACTIVITY LOG template."""
    # Load template (relative to project root). Users should keep MASTER ACTIVITY LOG.xlsx next to app folder.
    template_path = ACTIVITY_LOG_TEMPLATE
    if not os.path.exists(template_path):
        # Fallback: look one directory up from app.py (common unzip layout)
        alt = os.path.join(os.path.dirname(__file__), "MASTER ACTIVITY LOG.xlsx")
        if os.path.exists(alt):
            template_path = alt
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Master Activity Log"] if "Master Activity Log" in wb.sheetnames else wb.active

    # Header: month + case
    try:
        dt = datetime.fromisoformat(local_date)
    except Exception:
        dt = store_now()
    ws["A1"].value = f"MONTH:  {dt.strftime('%B').upper()}"
    # Case display name
    db = get_db()
    c = db.execute(
        "SELECT case_name FROM cases WHERE case_code = ? AND location_id = ?",
        (case_code, location_id),
    ).fetchone()
    case_name = (c["case_name"] if c else "") or ""
    ws["E1"].value = f"CASE #: {case_code} {case_name}".strip()

    # Pull history events for this case and date (store-local)
    # Include events where case is FROM or TO. Exclude counts.
    events = db.execute(
        """
        SELECT h.id, h.ts, h.action, h.upc, h.qty, h.from_case_code, h.to_case_code,
               h.trans_reg, h.dept_no, h.brief_desc, h.ticket_price, h.diamond_test,
               u.username,
               COALESCE(p.item_type,'') AS item_type,
               COALESCE(p.description,'') AS description
        FROM history h
        LEFT JOIN users u ON u.id = h.user_id
        LEFT JOIN products p ON p.upc = h.upc
        WHERE h.action IN ('RECEIVE','MOVE','SOLD','MISSING','RETURN')
          AND h.location_id = ?
          AND (h.from_case_code = ? OR h.to_case_code = ?)
        ORDER BY h.ts ASC, h.id ASC
        """,
        (location_id, case_code, case_code),
    ).fetchall()

    # Filter by local_date in Python to avoid SQLite timezone issues
    filtered = []
    for e in events:
        ld = _local_date_str_from_ts(e["ts"])
        if ld == local_date:
            filtered.append(e)

    # Fill rows starting at 10, maintaining merges B:C and D:E if present
    row = 10
    for e in filtered:
        ts = e["ts"]
        dt_local = _parse_iso_to_store(ts) or store_now()
        # DATE (Excel date-only)
        ws.cell(row, 1).value = datetime(dt_local.year, dt_local.month, dt_local.day)

        # ACTION
        action = (e["action"] or "").upper()

        # DOCUMENT # / TRANS/REG (use SYS-id)
        doc = (e["trans_reg"] or "").strip() if action in ("SOLD", "RETURN") else ""
        if not doc:
            doc = f"SYS-{e['id']}"
        ws.cell(row, 2).value = doc
        # B:C are merged in template; writing B is enough
        # DEPARTMENT # & BRIEF ITEM DESCRIPTION (D:E merged in template)

        if action == "SOLD":
            dept = (e["dept_no"] or "").strip()
            bdesc = (e["brief_desc"] or "").strip()
            ws.cell(row, 4).value = f"{dept} - {bdesc}".strip(" -")
        else:
            desc = ""
            if action == "RETURN":
                desc = (e["brief_desc"] or "").strip()
            if not desc:
                desc = (e["description"] or "").strip()
            if not desc:
                desc = (e["item_type"] or "").strip().upper() or "ITEM"
            if action == "MOVE":
                if e["to_case_code"] == case_code:
                    desc = f"FROM {e['from_case_code']} - {desc}"
                elif e["from_case_code"] == case_code:
                    desc = f"TO {e['to_case_code']} - {desc}"
            ws.cell(row, 4).value = desc

        # UPC
        ws.cell(row, 6).value = e["upc"]

        # TICKET PRICE
        if action == "SOLD":
            ws.cell(row, 7).value = e["ticket_price"]
        elif action == "RETURN":
            ws.cell(row, 7).value = e["ticket_price"]
        else:
            ws.cell(row, 7).value = None

        # DIA. TEST
        if action == "SOLD":
            ws.cell(row, 8).value = (e["diamond_test"] or "").strip().upper() or None
        elif action == "RECEIVE":
            ws.cell(row, 8).value = "NRT"
        elif action == "RETURN":
            ws.cell(row, 8).value = (e["diamond_test"] or "").strip().upper() or None
        else:
            ws.cell(row, 8).value = None

        # ITEM CODE
        ws.cell(row, 9).value = _item_code(e["item_type"])

        # REASON CODE
        ws.cell(row, 10).value = _reason_code(action)

        qty = int(e["qty"] or 0)
        qty_in = 0
        qty_out = 0
        if action == "RECEIVE":
            # goes into New Receipts (case_code may be NR)
            if e["to_case_code"] == case_code:
                qty_in = qty
            else:
                qty_out = qty
        elif action == "MOVE":
            if e["to_case_code"] == case_code:
                qty_in = qty
            elif e["from_case_code"] == case_code:
                qty_out = qty
        elif action == "RETURN":
            if e["to_case_code"] == case_code:
                qty_in = qty
            elif e["from_case_code"] == case_code:
                qty_out = qty
        else:
            # SOLD/MISSING remove from from_case_code
            if e["from_case_code"] == case_code:
                qty_out = qty

        ws.cell(row, 11).value = qty_in if qty_in else None
        ws.cell(row, 12).value = qty_out if qty_out else None

        # INITIALS (optional)
        ws.cell(row, 13).value = _initials_from_username(e["username"] or "")

        row += 1

    # Clear any leftover values below the filled region within the visible table area (optional)
    # Keep formatting intact.
    return wb

def build_daily_count_workbook(case_code: str, local_date: str, location_id: int):
    """Return an openpyxl workbook for the given case and local_date (YYYY-MM-DD) using Daily Count Sheet template."""
    template_path = DAILY_COUNT_TEMPLATE
    if not os.path.exists(template_path):
        alt = os.path.join(os.path.dirname(__file__), "Daily Count Sheet.xlsx")
        if os.path.exists(alt):
            template_path = alt

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    try:
        dt = datetime.fromisoformat(local_date)
    except Exception:
        dt = store_now()

    day_name = dt.strftime("%A").upper()
    date_display = dt.strftime("%m/%d/%Y")

    weekday_rows = {
        "SUNDAY": 5,
        "MONDAY": 15,
        "TUESDAY": 25,
        "WEDNESDAY": 35,
        "THURSDAY": 45,
        "FRIDAY": 55,
        "SATURDAY": 65,
    }
    start_row = weekday_rows.get(day_name, 5)

    ws[f"A{start_row}"].value = f"{day_name} - TODAY'S DATE:   {date_display}"
    ws[f"Q{start_row}"].value = f"CASE # {case_code}"

    db = get_db()
    previous_count = db.execute(
        """
        SELECT *
        FROM case_counts
        WHERE case_code=? AND local_date < ? AND location_id=?
        ORDER BY local_date DESC, id DESC
        LIMIT 1
        """,
        (case_code, local_date, location_id),
    ).fetchone()

    count = db.execute(
        """
        SELECT *
        FROM case_counts
        WHERE case_code=? AND local_date=? AND location_id=?
        ORDER BY id DESC
        LIMIT 1
        """,
        (case_code, local_date, location_id),
    ).fetchone()

    if previous_count:
        ws.cell(start_row + 3, 6).value = fmt_mmddyyyy(previous_count["local_date"])
        ws.cell(start_row + 5, 6).value = int(previous_count["total"])
        previous_initials = _initials_from_username(previous_count["username"] or "")
        if previous_initials:
            ws.cell(start_row + 7, 6).value = previous_initials

    total_in, total_out, net = _daily_activity_totals(case_code, local_date, location_id)
    ws.cell(start_row + 3, 13).value = total_in
    ws.cell(start_row + 3, 14).value = total_out
    ws.cell(start_row + 3, 15).value = net

    if count:
        row_map = {
            "necklaces": start_row + 3,
            "earrings": start_row + 4,
            "rings": start_row + 5,
            "bracelets": start_row + 6,
            "other": start_row + 7,
        }
        for key, row in row_map.items():
            case_value = int(count[key])
            reserve_value = int(count[f"reserve_{key}"])
            total_value = case_value + reserve_value
            ws.cell(row, 10).value = case_value
            ws.cell(row, 11).value = reserve_value
            ws.cell(row, 12).value = total_value

        total = int(count["total"])
        ws.cell(start_row + 8, 12).value = total
        ws.cell(start_row + 5, 16).value = total

        initials = _initials_from_username(count["username"] or "")
        if initials:
            ws.cell(start_row + 8, 10).value = initials
            ws.cell(start_row + 8, 16).value = initials

        ws.cell(start_row + 6, 16).value = net
        ws.cell(start_row + 7, 16).value = total - net

        notes = (count["notes"] or "").strip()
        if notes:
            ws.cell(start_row + 3, 20).value = notes

    return wb

@app.route("/counts")
@login_required
def counts():
    db = get_db()
    today = local_date_str()
    location_id = current_location_id()

    cases = db.execute(
        f"""
        SELECT c.case_code, c.case_name, c.is_virtual, c.is_active,
               COALESCE(SUM(i.qty), 0) AS total_qty,
               COALESCE(COUNT(DISTINCT i.upc), 0) AS distinct_upcs
        FROM cases c
        LEFT JOIN inventory i ON i.case_code = c.case_code AND i.location_id = c.location_id
        WHERE c.is_active = 1 AND c.location_id = ?
        GROUP BY c.case_code
        {case_order_sql()}
        """,
        (location_id,),
    ).fetchall()

    # Latest count per case for today
    counts_rows = db.execute(
        """
        SELECT cc.*
        FROM case_counts cc
        JOIN (
          SELECT case_code, MAX(id) AS max_id
          FROM case_counts
          WHERE local_date=? AND location_id=?
          GROUP BY case_code
        ) m ON m.max_id = cc.id
        """,
        (today, location_id),
    ).fetchall()
    counts_map = {r["case_code"]: r for r in counts_rows}

    sys_totals = {
        c["case_code"]: {
            "case": case_type_totals(c["case_code"], LOCATION_CASE, location_id),
            "reserve": case_type_totals(c["case_code"], LOCATION_RESERVE, location_id),
            "combined": case_type_totals(c["case_code"], location_id=location_id),
        }
        for c in cases
    }

    return render_template(
        "counts.html",
        cases=cases,
        counts_map=counts_map,
        sys_totals=sys_totals,
        today=today,
        count_categories=COUNT_CATEGORIES,
        user=current_user(),
        new_receipts_code=new_receipts_case_code(location_id),
    )


@app.route("/counts/<case_code>", methods=["GET", "POST"])
@login_required
def count_case(case_code: str):
    case_code = (case_code or "").strip()
    if not ensure_case_exists(case_code):
        flash("Case not found.", "danger")
        return redirect(url_for("counts"))

    db = get_db()
    location_id = current_location_id()
    case = db.execute(
        "SELECT * FROM cases WHERE case_code=? AND location_id=?",
        (case_code, location_id),
    ).fetchone()
    if not case or case["is_active"] != 1:
        flash("Case not found.", "danger")
        return redirect(url_for("counts"))

    today = local_date_str()
    sys_case = case_type_totals(case_code, LOCATION_CASE, location_id)
    sys_reserve = case_type_totals(case_code, LOCATION_RESERVE, location_id)
    sys_combined = case_type_totals(case_code, location_id=location_id)

    last_count = db.execute(
        """
        SELECT *
        FROM case_counts
        WHERE case_code=? AND local_date=? AND location_id=?
        ORDER BY id DESC
        LIMIT 1
        """,
        (case_code, today, location_id),
    ).fetchone()

    if request.method == "POST":
        def to_int(name: str) -> int:
            try:
                return int((request.form.get(name) or "0").strip())
            except ValueError:
                return -1

        counts = {field: to_int(field) for field in COUNT_FIELDS}
        reserve_counts = {field: to_int(field) for field in RESERVE_COUNT_FIELDS}
        notes = (request.form.get("notes") or "").strip() or None

        if any(value < 0 for value in counts.values()) or any(value < 0 for value in reserve_counts.values()):
            flash("Counts must be whole numbers (0 or higher).", "danger")
            return redirect(url_for("count_case", case_code=case_code))

        total = sum(counts.values()) + sum(reserve_counts.values())
        u = current_user()

        db.execute(
            """
            INSERT INTO case_counts
              (ts_utc, local_date, case_code, location_id, user_id, username,
               bracelets, rings, earrings, necklaces, other,
               reserve_bracelets, reserve_rings, reserve_earrings, reserve_necklaces, reserve_other,
               total, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (utc_now(), today, case_code, location_id,
             u.id if u else None, u.username if u else None,
             counts["bracelets"], counts["rings"], counts["earrings"], counts["necklaces"], counts["other"],
             reserve_counts["reserve_bracelets"], reserve_counts["reserve_rings"], reserve_counts["reserve_earrings"],
             reserve_counts["reserve_necklaces"], reserve_counts["reserve_other"],
             total, notes),
        )
        db.commit()

        flash(f"Count saved for Case {case_code} ({today}).", "success")
        return redirect(url_for("view_case", case_code=case_code))

    return render_template(
        "count_case.html",
        case=case,
        today=today,
        sys_case=sys_case,
        sys_reserve=sys_reserve,
        sys_combined=sys_combined,
        last_count=last_count,
        count_categories=COUNT_CATEGORIES,
        user=current_user(),
    )


# ---------------- History ----------------
def log_history(
    action: str,
    upc: Optional[str] = None,
    qty: Optional[int] = None,
    from_case_code: Optional[str] = None,
    to_case_code: Optional[str] = None,
    notes: Optional[str] = None,
    trans_reg: Optional[str] = None,
    dept_no: Optional[str] = None,
    brief_desc: Optional[str] = None,
    ticket_price: Optional[float] = None,
    diamond_test: Optional[str] = None,
):
    db = get_db()
    u = current_user()
    location_id = current_location_id()
    db.execute(
        """
        INSERT INTO history (ts, user_id, username, action, upc, qty, from_case_code, to_case_code, notes, trans_reg, dept_no, brief_desc, ticket_price, diamond_test, location_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            utc_now(),
            u.id if u else None,
            u.username if u else None,
            action,
            upc,
            qty,
            from_case_code,
            to_case_code,
            notes,
            trans_reg,
            dept_no,
            brief_desc,
            ticket_price,
            diamond_test,
            location_id,
        ),
    )
    db.commit()


# ---------------- Parsing: UPC lists w/ optional qty ----------------
def parse_upc_lines(text: str) -> Dict[str, int]:
    """
    Accepts:
      - one UPC per line
      - or "UPC,qty"
    Returns aggregated counts by UPC.
    """
    out: Dict[str, int] = {}
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line:
            continue

        if "," in line:
            left, right = line.split(",", 1)
            upc = left.strip()
            try:
                qty = int(right.strip())
            except ValueError:
                qty = 0
        else:
            upc = line
            qty = 1

        upc = upc.strip()
        if not upc or qty <= 0:
            continue
        out[upc] = out.get(upc, 0) + qty
    return out


# ---------------- Inventory operations ----------------
def ensure_case_exists(case_code: str) -> bool:
    db = get_db()
    location_id = current_location_id()
    row = db.execute(
        "SELECT case_code FROM cases WHERE case_code=? AND is_active=1 AND location_id=?",
        (case_code, location_id),
    ).fetchone()
    return bool(row)


def upsert_product(upc: str, description: Optional[str], item_type: Optional[str] = None):
    db = get_db()

    if item_type and item_type not in ALLOWED_ITEM_TYPES:
        item_type = None

    existing = db.execute("SELECT description, item_type FROM products WHERE upc=?", (upc,)).fetchone()
    if existing:
        updates = {}
        if description and not existing["description"]:
            updates["description"] = description
        if item_type and (not existing["item_type"]):
            updates["item_type"] = item_type

        if updates:
            sets = ", ".join([f"{k}=?" for k in updates.keys()])
            params = list(updates.values()) + [upc]
            db.execute(f"UPDATE products SET {sets} WHERE upc=?", params)
    else:
        db.execute(
            "INSERT INTO products (upc, description, item_type) VALUES (?, ?, ?)",
            (upc, description, item_type),
        )


def add_qty(
    case_code: str,
    upc: str,
    qty: int,
    location: str = LOCATION_CASE,
    location_id: Optional[int] = None,
):
    db = get_db()
    location_id = location_id or current_location_id()
    if not location_id:
        raise ValueError("location_id is required for inventory updates")
    if location not in INVENTORY_LOCATIONS:
        location = LOCATION_CASE
    db.execute(
        """
        INSERT INTO inventory (case_code, location_id, upc, location, qty)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(case_code, location_id, upc, location) DO UPDATE SET qty = qty + excluded.qty
        """,
        (case_code, location_id, upc, location, qty),
    )


def remove_qty(
    case_code: str,
    upc: str,
    qty: int,
    location: str = LOCATION_CASE,
    location_id: Optional[int] = None,
) -> Tuple[bool, int]:
    db = get_db()
    location_id = location_id or current_location_id()
    if not location_id:
        raise ValueError("location_id is required for inventory updates")
    if location not in INVENTORY_LOCATIONS:
        location = LOCATION_CASE
    row = db.execute(
        "SELECT qty FROM inventory WHERE case_code=? AND location_id=? AND upc=? AND location=?",
        (case_code, location_id, upc, location),
    ).fetchone()
    if not row:
        return False, 0
    have = int(row["qty"])
    if have < qty:
        return False, have

    new_qty = have - qty
    if new_qty == 0:
        db.execute(
            "DELETE FROM inventory WHERE case_code=? AND location_id=? AND upc=? AND location=?",
            (case_code, location_id, upc, location),
        )
    else:
        db.execute(
            "UPDATE inventory SET qty=? WHERE case_code=? AND location_id=? AND upc=? AND location=?",
            (new_qty, case_code, location_id, upc, location),
        )
    return True, new_qty



def case_order_sql() -> str:
    # numeric sort for '01'..'30'
    # expects cases table is aliased as 'c'
    return """
      ORDER BY
        c.is_virtual DESC,
        CASE
          WHEN c.case_code GLOB '[0-9]*' THEN CAST(c.case_code AS INTEGER)
          ELSE 999999
        END,
        c.case_code
    """
def _validate_have_qty(
    case_code: str,
    upc_map: Dict[str, int],
    location: str = LOCATION_CASE,
    location_id: Optional[int] = None,
) -> list[str]:
    db = get_db()
    location_id = location_id or current_location_id()
    if not location_id:
        raise ValueError("location_id is required for inventory updates")
    if location not in INVENTORY_LOCATIONS:
        location = LOCATION_CASE
    problems = []
    for upc, need in upc_map.items():
        row = db.execute(
            "SELECT qty FROM inventory WHERE case_code=? AND location_id=? AND upc=? AND location=?",
            (case_code, location_id, upc, location),
        ).fetchone()
        have = int(row["qty"]) if row else 0
        if have < need:
            problems.append(f"{upc}: need {need}, have {have}")
    return problems


def case_type_totals(
    case_code: str,
    location: Optional[str] = None,
    location_id: Optional[int] = None,
) -> dict:
    """Compute live totals for a case, grouped by item_type."""
    db = get_db()
    location_id = location_id or current_location_id()
    if not location_id:
        raise ValueError("location_id is required for inventory lookups")
    pieces = []
    params = []
    for category in ITEM_CATEGORIES:
        pieces.append(
            f"COALESCE(SUM(CASE WHEN p.item_type=? THEN inv.qty ELSE 0 END),0) AS {category['count_key']}"
        )
        params.append(category["name"])
    pieces.append("COALESCE(SUM(inv.qty),0) AS total")
    pieces.append("COALESCE(SUM(CASE WHEN p.item_type IS NULL OR p.item_type='' THEN inv.qty ELSE 0 END),0) AS unknown")
    sql = f"""
        SELECT
          {", ".join(pieces)}
        FROM inventory inv
        LEFT JOIN products p ON p.upc = inv.upc
        WHERE inv.case_code = ? AND inv.location_id = ?
        """
    params.append(case_code)
    params.append(location_id)
    if location in INVENTORY_LOCATIONS:
        sql += " AND inv.location = ?"
        params.append(location)
    row = db.execute(sql, params).fetchone()
    base = {c["count_key"]: 0 for c in ITEM_CATEGORIES}
    base.update({"total": 0, "unknown": 0})
    return dict(row) if row else base



# ---------------- Setup/Login ----------------
@app.route("/setup", methods=["GET", "POST"])
def setup():
    db = get_db()
    existing = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
    if existing > 0:
        flash("Setup is already complete.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        if not username or not password or len(password) < 8:
            flash("Username required. Password must be at least 8 characters.", "danger")
            return redirect(url_for("setup"))

        default_location = db.execute(
            "SELECT id FROM locations ORDER BY id LIMIT 1"
        ).fetchone()
        default_location_id = default_location["id"] if default_location else None
        db.execute(
            """
            INSERT INTO users (username, password_hash, role, is_active, location_id, created_at)
            VALUES (?,?,?,?,?,?)
            """,
            (username, generate_password_hash(password), ROLE_ADMIN, 1, default_location_id, utc_now()),
        )
        db.commit()
        flash("Admin user created. Log in.", "success")
        return redirect(url_for("login"))

    return render_template("setup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if not is_setup_complete():
        return redirect(url_for("setup"))

    if current_user():
        if current_user().role == ROLE_ADMIN and not current_location_id():
            return redirect(url_for("select_location"))
        return redirect(url_for("index"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        db = get_db()
        row = db.execute(
            "SELECT id, username, password_hash, role, is_active, location_id FROM users WHERE username=?",
            (username,),
        ).fetchone()
        if not row or row["is_active"] != 1 or not check_password_hash(row["password_hash"], password):
            flash("Invalid login.", "danger")
            return redirect(url_for("login"))

        session["user_id"] = row["id"]
        if row["role"] == ROLE_STAFF:
            session["location_id"] = row["location_id"]
            flash("Logged in.", "success")
            nxt = request.args.get("next")
            return redirect(nxt or url_for("index"))

        session.pop("location_id", None)
        flash("Logged in.", "success")
        return redirect(url_for("select_location"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    session.pop("location_id", None)
    flash("Logged out.", "success")
    return redirect(url_for("login"))


# ---------------- Locations ----------------
@app.route("/locations/select", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def select_location():
    db = get_db()
    locations = db.execute(
        "SELECT id, name FROM locations WHERE is_active=1 ORDER BY name"
    ).fetchall()

    if request.method == "POST":
        location_id = request.form.get("location_id")
        try:
            location_id_int = int(location_id)
        except (TypeError, ValueError):
            flash("Select a valid location.", "danger")
            return redirect(url_for("select_location"))

        exists = db.execute(
            "SELECT id FROM locations WHERE id=? AND is_active=1",
            (location_id_int,),
        ).fetchone()
        if not exists:
            flash("Select a valid location.", "danger")
            return redirect(url_for("select_location"))

        session["location_id"] = location_id_int
        flash("Location selected.", "success")
        return redirect(url_for("index"))

    return render_template(
        "select_location.html",
        locations=locations,
        current_location_id=session.get("location_id"),
        user=current_user(),
    )


@app.route("/admin/locations", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def locations_admin():
    db = get_db()
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if not name:
            flash("Location name is required.", "danger")
            return redirect(url_for("locations_admin"))
        existing = db.execute(
            "SELECT id FROM locations WHERE name=?",
            (name,),
        ).fetchone()
        if existing:
            flash("That location already exists.", "danger")
            return redirect(url_for("locations_admin"))

        db.execute(
            "INSERT INTO locations (name, is_active, created_at) VALUES (?, 1, ?)",
            (name, utc_now()),
        )
        location_id = db.execute(
            "SELECT id FROM locations WHERE name=?",
            (name,),
        ).fetchone()["id"]
        new_receipts_code = new_receipts_case_code(location_id)
        db.execute(
            """
            INSERT OR IGNORE INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at)
            VALUES (?, ?, ?, 1, 1, ?)
            """,
            (new_receipts_code, location_id, NEW_RECEIPTS_NAME, utc_now()),
        )
        returns_code = returns_case_code(location_id)
        db.execute(
            """
            INSERT OR IGNORE INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at)
            VALUES (?, ?, ?, 1, 1, ?)
            """,
            (returns_code, location_id, RETURNS_NAME, utc_now()),
        )
        db.commit()
        flash("Location created.", "success")
        return redirect(url_for("locations_admin"))

    locations = db.execute(
        "SELECT id, name, is_active, created_at FROM locations ORDER BY name"
    ).fetchall()
    return render_template(
        "locations.html",
        locations=locations,
        user=current_user(),
    )

# ---------------- Main: Case Grid ----------------
@app.route("/")
@login_required
def index():
    db = get_db()
    location_id = current_location_id()
    cases = db.execute(
        f"""
        SELECT c.case_code, c.case_name, c.is_virtual, c.is_active,
               COALESCE(SUM(i.qty), 0) AS total_qty,
               COALESCE(COUNT(DISTINCT i.upc), 0) AS distinct_upcs
        FROM cases c
        LEFT JOIN inventory i ON i.case_code = c.case_code AND i.location_id = c.location_id
        WHERE c.is_active = 1 AND c.location_id = ?
        GROUP BY c.case_code
        {case_order_sql()}
        """,
        (location_id,),
    ).fetchall()

    recent = db.execute(
        """
        SELECT h.*, p.item_type
        FROM history h
        LEFT JOIN products p ON p.upc = h.upc
        WHERE h.location_id = ?
        ORDER BY h.id DESC
        LIMIT 25
        """,
        (location_id,),
    ).fetchall()

    return render_template("index.html", cases=cases, recent=recent, user=current_user())


@app.route("/cases/new", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def create_case():
    case_code = (request.form.get("case_code") or "").strip()
    case_name = (request.form.get("case_name") or "").strip()
    location_id = current_location_id()

    if not case_code or not case_name:
        flash("Case number and name are required.", "danger")
        return redirect(url_for("index"))
    if case_code == new_receipts_case_code(location_id):
        flash("That case code is reserved.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    try:
        db.execute(
            "INSERT INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at) VALUES (?,?,?,?,?,?)",
            (case_code, location_id, case_name, 0, 1, utc_now()),
        )
        db.commit()
        log_history(ACTION_CASE_CREATE, notes=f"Created case {case_code} ({case_name})")
        flash(f"Case {case_code} created.", "success")
    except sqlite3.IntegrityError:
        flash("That case number already exists.", "danger")

    return redirect(url_for("index"))


@app.route("/cases/<case_code>")
@login_required
def view_case(case_code: str):
    case_code = (case_code or "").strip()
    db = get_db()
    location_id = current_location_id()

    case = db.execute(
        "SELECT * FROM cases WHERE case_code=? AND location_id=?",
        (case_code, location_id),
    ).fetchone()
    if not case or case["is_active"] != 1:
        flash("Case not found.", "danger")
        return redirect(url_for("index"))

    case_items = db.execute(
        """
        SELECT inv.upc, inv.qty, p.description, p.item_type
        FROM inventory inv
        LEFT JOIN products p ON p.upc = inv.upc
        WHERE inv.case_code = ? AND inv.location_id = ? AND inv.location = ?
        ORDER BY inv.upc
        """,
        (case_code, location_id, LOCATION_CASE),
    ).fetchall()

    reserve_items = db.execute(
        """
        SELECT inv.upc, inv.qty, p.description, p.item_type
        FROM inventory inv
        LEFT JOIN products p ON p.upc = inv.upc
        WHERE inv.case_code = ? AND inv.location_id = ? AND inv.location = ?
        ORDER BY inv.upc
        """,
        (case_code, location_id, LOCATION_RESERVE),
    ).fetchall()

    totals = db.execute(
        """
        SELECT COALESCE(SUM(qty),0) AS total_qty, COALESCE(COUNT(DISTINCT upc),0) AS distinct_upcs
        FROM inventory
        WHERE case_code=? AND location_id=?
        """,
        (case_code, location_id),
    ).fetchone()

    case_totals = db.execute(
        """
        SELECT COALESCE(SUM(qty),0) AS total_qty, COALESCE(COUNT(DISTINCT upc),0) AS distinct_upcs
        FROM inventory
        WHERE case_code=? AND location_id=? AND location=?
        """,
        (case_code, location_id, LOCATION_CASE),
    ).fetchone()

    reserve_totals = db.execute(
        """
        SELECT COALESCE(SUM(qty),0) AS total_qty, COALESCE(COUNT(DISTINCT upc),0) AS distinct_upcs
        FROM inventory
        WHERE case_code=? AND location_id=? AND location=?
        """,
        (case_code, location_id, LOCATION_RESERVE),
    ).fetchone()

    case_type_totals_data = case_type_totals(case_code, LOCATION_CASE, location_id)
    reserve_type_totals = case_type_totals(case_code, LOCATION_RESERVE, location_id)
    combined_type_totals = case_type_totals(case_code, location_id=location_id)

    # Latest physical count for today (store-local)
    today = local_date_str()
    last_count = db.execute(
        """
        SELECT *
        FROM case_counts
        WHERE case_code=? AND local_date=? AND location_id=?
        ORDER BY id DESC
        LIMIT 1
        """,
        (case_code, today, location_id),
    ).fetchone()


    history_rows = db.execute(
        """
        SELECT h.*, p.item_type
        FROM history h
        LEFT JOIN products p ON p.upc = h.upc
        WHERE h.location_id = ? AND (h.from_case_code = ? OR h.to_case_code = ?)
        ORDER BY id DESC
        LIMIT 150
        """,
        (location_id, case_code, case_code),
    ).fetchall()

    active_cases = db.execute(
        f"""
        SELECT c.case_code, c.case_name, c.is_virtual
        FROM cases c
        WHERE c.is_active=1 AND c.location_id=?
        {case_order_sql()}
        """,
        (location_id,),
    ).fetchall()

    return render_template(
        "case.html",
        case=case,
        items=case_items,
        reserve_items=reserve_items,
        totals=totals,
        case_totals=case_totals,
        reserve_totals=reserve_totals,
        case_type_totals=case_type_totals_data,
        reserve_type_totals=reserve_type_totals,
        combined_type_totals=combined_type_totals,
        last_count=last_count,
        history=history_rows,
        active_cases=active_cases,
        count_categories=COUNT_CATEGORIES,
        user=current_user(),
        new_receipts_code=new_receipts_case_code(location_id),
    )


@app.get("/api/case/<case_code>/items")
@login_required
def api_case_items(case_code: str):
    case_code = (case_code or "").strip()
    if not ensure_case_exists(case_code):
        return jsonify({"ok": False, "error": "Case not found"}), 404

    location = (request.args.get("location") or "").strip().upper()
    if location not in INVENTORY_LOCATIONS:
        location = LOCATION_CASE
    location_id = current_location_id()

    db = get_db()
    rows = db.execute(
        """
        SELECT inv.upc, inv.qty, p.description, p.item_type
        FROM inventory inv
        LEFT JOIN products p ON p.upc = inv.upc
        WHERE inv.case_code = ? AND inv.location_id = ? AND inv.location = ?
        ORDER BY inv.upc
        """,
        (case_code, location_id, location),
    ).fetchall()

    items = [
        {
            "upc": r["upc"],
            "qty": int(r["qty"]),
            "description": r["description"] or "",
            "item_type": r["item_type"] or "",
        }
        for r in rows
    ]
    return jsonify({"ok": True, "case_code": case_code, "location": location, "items": items})


@app.get("/cases/<case_code>/edit")
@login_required
@role_required(ROLE_ADMIN)
def edit_case(case_code: str):
    case_code = (case_code or "").strip()
    db = get_db()
    location_id = current_location_id()

    case = db.execute(
        "SELECT * FROM cases WHERE case_code=? AND location_id=?",
        (case_code, location_id),
    ).fetchone()
    if not case or case["is_active"] != 1:
        flash("Case not found.", "danger")
        return redirect(url_for("index"))

    return render_template("edit_case.html", case=case, user=current_user())


@app.post("/cases/<case_code>/edit")
@login_required
@role_required(ROLE_ADMIN)
def edit_case_post(case_code: str):
    case_code = (case_code or "").strip()
    new_name = (request.form.get("case_name") or "").strip()
    location_id = current_location_id()

    if not new_name:
        flash("Case name is required.", "danger")
        return redirect(url_for("edit_case", case_code=case_code))

    db = get_db()
    case = db.execute(
        "SELECT * FROM cases WHERE case_code=? AND location_id=?",
        (case_code, location_id),
    ).fetchone()
    if not case or case["is_active"] != 1:
        flash("Case not found.", "danger")
        return redirect(url_for("index"))

    old_name = case["case_name"]
    if old_name == new_name:
        flash("No changes made.", "warning")
        return redirect(url_for("view_case", case_code=case_code))

    db.execute(
        "UPDATE cases SET case_name=? WHERE case_code=? AND location_id=?",
        (new_name, case_code, location_id),
    )
    db.commit()

    log_history(ACTION_CASE_EDIT, notes=f"Renamed case {case_code}: '{old_name}' → '{new_name}'")
    flash(f"Case {case_code} renamed.", "success")
    return redirect(url_for("view_case", case_code=case_code))


@app.route("/cases/<case_code>/delete", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def delete_case(case_code: str):
    case_code = (case_code or "").strip()
    location_id = current_location_id()
    if case_code == new_receipts_case_code(location_id):
        flash("You can’t delete New Receipts.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cnt = db.execute(
        """
        SELECT COALESCE(SUM(i.qty),0) AS t
        FROM inventory i
        JOIN cases c ON c.case_code = i.case_code AND c.location_id = i.location_id
        WHERE i.case_code=? AND i.location_id=?
        """,
        (case_code, location_id),
    ).fetchone()["t"]
    if cnt and int(cnt) > 0:
        flash("Case must be empty before deletion. Move items out first.", "danger")
        return redirect(url_for("view_case", case_code=case_code))

    db.execute(
        "UPDATE cases SET is_active=0 WHERE case_code=? AND location_id=?",
        (case_code, location_id),
    )
    db.commit()
    log_history(ACTION_CASE_DELETE, notes=f"Deleted/archived case {case_code}")
    flash(f"Case {case_code} deleted (archived).", "success")
    return redirect(url_for("index"))


# ---------------- Workbench actions (on case page) ----------------
@app.post("/cases/<case_code>/move_out")
@login_required
def move_out_of_case(case_code: str):
    case_code = (case_code or "").strip()
    to_case = (request.form.get("to_case_code") or "").strip()
    description = (request.form.get("description") or "").strip() or None
    upc_map = parse_upc_lines(request.form.get("upcs", ""))
    if not ensure_case_exists(case_code):
        flash("Case not found.", "danger")
        return redirect(url_for("index"))
    if not to_case or not ensure_case_exists(to_case):
        flash("Destination case not found.", "danger")
        return redirect(url_for("view_case", case_code=case_code))
    if to_case == case_code:
        flash("Destination can’t be the same case.", "danger")
        return redirect(url_for("view_case", case_code=case_code))
    if not upc_map:
        flash("Scan/enter at least one UPC to move.", "danger")
        return redirect(url_for("view_case", case_code=case_code))

    problems = _validate_have_qty(case_code, upc_map, LOCATION_CASE)
    if problems:
        flash("Not enough qty to move for: " + "; ".join(problems), "danger")
        return redirect(url_for("view_case", case_code=case_code))

    db = get_db()
    for upc, qty in upc_map.items():
        upsert_product(upc, description, item_type=None)
        ok, _ = remove_qty(case_code, upc, qty, LOCATION_CASE)
        if ok:
            add_qty(to_case, upc, qty, LOCATION_CASE)
            log_history(ACTION_MOVE, upc=upc, qty=qty, from_case_code=case_code, to_case_code=to_case, notes="Moved from case workbench")
    db.commit()

    flash(f"Moved {sum(upc_map.values())} unit(s) from {case_code} → {to_case}.", "success")
    return redirect(url_for("view_case", case_code=to_case))


@app.post("/cases/<case_code>/transfer_location")
@login_required
def transfer_case_location(case_code: str):
    case_code = (case_code or "").strip()
    from_location = (request.form.get("from_location") or "").strip().upper()
    to_location = (request.form.get("to_location") or "").strip().upper()
    upc_map = parse_upc_lines(request.form.get("upcs", ""))

    if not ensure_case_exists(case_code):
        flash("Case not found.", "danger")
        return redirect(url_for("index"))
    if from_location not in INVENTORY_LOCATIONS or to_location not in INVENTORY_LOCATIONS:
        flash("Choose Case or Reserve for the move.", "danger")
        return redirect(url_for("view_case", case_code=case_code))
    if from_location == to_location:
        flash("From and To locations can’t be the same.", "danger")
        return redirect(url_for("view_case", case_code=case_code))
    if not upc_map:
        flash("Scan/enter at least one UPC to move.", "danger")
        return redirect(url_for("view_case", case_code=case_code))

    problems = _validate_have_qty(case_code, upc_map, from_location)
    if problems:
        flash("Not enough qty to move for: " + "; ".join(problems), "danger")
        return redirect(url_for("view_case", case_code=case_code))

    db = get_db()
    for upc, qty in upc_map.items():
        ok, _ = remove_qty(case_code, upc, qty, from_location)
        if ok:
            add_qty(case_code, upc, qty, to_location)
            log_history(
                ACTION_MOVE,
                upc=upc,
                qty=qty,
                from_case_code=case_code,
                to_case_code=case_code,
                notes=f"Moved {from_location} → {to_location} (case workbench)",
            )
    db.commit()

    flash(f"Moved {sum(upc_map.values())} unit(s) {from_location} → {to_location} for case {case_code}.", "success")
    return redirect(url_for("view_case", case_code=case_code))


def parse_sold_fields(form: dict) -> Tuple[Optional[dict], Optional[str]]:
    trans_reg = (form.get("trans_reg") or "").strip()
    dept_no = (form.get("dept_no") or "").strip()
    brief_desc = (form.get("brief_desc") or "").strip()
    ticket_price_raw = (form.get("ticket_price") or "").strip()
    diamond_test = (form.get("diamond_test") or "").strip().upper()

    if (not trans_reg) or (not dept_no) or (not brief_desc) or (not ticket_price_raw) or (diamond_test not in DIAMOND_TEST_OPTIONS):
        return None, "For SOLD you must enter Transaction/Register #, Department #, Brief Description, Ticket Price, and Diamond Test (Y/N/NRT)."

    try:
        ticket_price = float(ticket_price_raw.replace("$", "").replace(",", ""))
    except Exception:
        return None, "Ticket Price must be a valid number (example: 199.99)."

    return {
        "trans_reg": trans_reg,
        "dept_no": dept_no,
        "brief_desc": brief_desc,
        "ticket_price": ticket_price,
        "diamond_test": diamond_test,
    }, None


@app.post("/cases/<case_code>/sell_out")
@login_required
def sell_from_case(case_code: str):
    case_code = (case_code or "").strip()
    upc_map = parse_upc_lines(request.form.get("upcs", ""))

    if not ensure_case_exists(case_code):
        flash("Case not found.", "danger")
        return redirect(url_for("index"))
    if not upc_map:
        flash("Scan/enter at least one UPC to sell.", "danger")
        return redirect(url_for("view_case", case_code=case_code))

    sold_fields, error = parse_sold_fields(request.form)
    if error:
        flash(error, "danger")
        return redirect(url_for("view_case", case_code=case_code))

    problems = _validate_have_qty(case_code, upc_map, LOCATION_CASE)
    if problems:
        flash("Not enough qty to sell for: " + "; ".join(problems), "danger")
        return redirect(url_for("view_case", case_code=case_code))

    db = get_db()
    for upc, qty in upc_map.items():
        ok, _ = remove_qty(case_code, upc, qty, LOCATION_CASE)
        if ok:
            log_history(
                ACTION_SOLD,
                upc=upc,
                qty=qty,
                from_case_code=case_code,
                notes="Sold from case workbench",
                trans_reg=sold_fields["trans_reg"],
                dept_no=sold_fields["dept_no"],
                brief_desc=sold_fields["brief_desc"],
                ticket_price=sold_fields["ticket_price"],
                diamond_test=sold_fields["diamond_test"],
            )
    db.commit()

    flash(f"Sold {sum(upc_map.values())} unit(s) from case {case_code}.", "success")
    return redirect(url_for("view_case", case_code=case_code))


@app.post("/cases/<case_code>/missing_out")
@login_required
def missing_from_case(case_code: str):
    case_code = (case_code or "").strip()
    upc_map = parse_upc_lines(request.form.get("upcs", ""))
    notes = (request.form.get("notes") or "").strip() or None

    if not ensure_case_exists(case_code):
        flash("Case not found.", "danger")
        return redirect(url_for("index"))
    if not upc_map:
        flash("Scan/enter at least one UPC to mark missing.", "danger")
        return redirect(url_for("view_case", case_code=case_code))

    problems = _validate_have_qty(case_code, upc_map, LOCATION_CASE)
    if problems:
        flash("Not enough qty to mark missing for: " + "; ".join(problems), "danger")
        return redirect(url_for("view_case", case_code=case_code))

    db = get_db()
    for upc, qty in upc_map.items():
        ok, _ = remove_qty(case_code, upc, qty, LOCATION_CASE)
        if ok:
            log_history(ACTION_MISSING, upc=upc, qty=qty, from_case_code=case_code, notes=notes or "Marked missing from case workbench")
    db.commit()

    flash(f"Marked MISSING: {sum(upc_map.values())} unit(s) from case {case_code}.", "success")
    return redirect(url_for("view_case", case_code=case_code))


# ---------------- Receive (required item type) ----------------
@app.route("/receive", methods=["GET", "POST"])
@login_required
def receive():
    location_id = current_location_id()
    new_receipts_code = new_receipts_case_code(location_id)
    if request.method == "POST":
        description = (request.form.get("description") or "").strip() or None
        item_type = (request.form.get("item_type") or "").strip()
        upc_map = parse_upc_lines(request.form.get("upcs", ""))

        if item_type not in ALLOWED_ITEM_TYPES:
            flash(f"Select an item type ({', '.join(ITEM_TYPES_ORDER)}).", "danger")
            return redirect(url_for("receive"))

        if not upc_map:
            flash("Scan/enter at least one UPC.", "danger")
            return redirect(url_for("receive"))

        db = get_db()
        for upc, qty in upc_map.items():
            upsert_product(upc, description, item_type=item_type)
            add_qty(new_receipts_code, upc, qty, LOCATION_CASE)
            log_history(
                ACTION_RECEIVE,
                upc=upc,
                qty=qty,
                to_case_code=new_receipts_code,
                notes=f"Received into New Receipts ({item_type})",
            )
        db.commit()

        flash(f"Received {sum(upc_map.values())} unit(s) into New Receipts.", "success")
        return redirect(url_for("view_case", case_code=new_receipts_code))

    return render_template("receive.html", user=current_user(), item_types=ITEM_TYPES_ORDER)


# ---------------- Returns ----------------
@app.route("/returns", methods=["GET", "POST"])
@login_required
def returns():
    location_id = current_location_id()
    returns_code = returns_case_code(location_id)
    if request.method == "POST":
        trans_reg = (request.form.get("return_trans") or "").strip()
        upc = (request.form.get("upc") or "").strip()
        item_type = (request.form.get("item_type") or "").strip()
        price_raw = (request.form.get("price") or "").strip()
        diamond_test = (request.form.get("diamond_test") or "").strip().upper()
        description = (request.form.get("description") or "").strip()

        if not trans_reg or not upc or not price_raw or not description or not diamond_test:
            flash("Return transaction #, UPC, price, diamond test, and description are required.", "danger")
            return redirect(url_for("returns"))
        if item_type not in ALLOWED_ITEM_TYPES:
            flash(f"Select an item type ({', '.join(ITEM_TYPES_ORDER)}).", "danger")
            return redirect(url_for("returns"))
        if diamond_test not in RETURN_DIAMOND_OPTIONS:
            flash("Select a diamond test value (Y, N, or N/A).", "danger")
            return redirect(url_for("returns"))

        try:
            price = float(price_raw.replace("$", "").replace(",", ""))
        except Exception:
            flash("Price must be a valid number (example: 199.99).", "danger")
            return redirect(url_for("returns"))

        if not ensure_case_exists(returns_code):
            db = get_db()
            db.execute(
                """
                INSERT OR IGNORE INTO cases (case_code, location_id, case_name, is_virtual, is_active, created_at)
                VALUES (?, ?, ?, 1, 1, ?)
                """,
                (returns_code, location_id, RETURNS_NAME, utc_now()),
            )
            db.commit()

        upsert_product(upc, description, item_type=item_type)
        add_qty(returns_code, upc, 1, LOCATION_CASE)
        log_history(
            ACTION_RETURN,
            upc=upc,
            qty=1,
            to_case_code=returns_code,
            notes="Return intake",
            trans_reg=trans_reg,
            brief_desc=description,
            ticket_price=price,
            diamond_test=diamond_test,
        )

        flash(f"Return added to {returns_code}.", "success")
        return redirect(url_for("view_case", case_code=returns_code))

    return render_template("returns.html", user=current_user(), item_types=ITEM_TYPES_ORDER)


# ---------------- Bulk Move page ----------------
@app.route("/move", methods=["GET", "POST"])
@login_required
def move():
    db = get_db()
    location_id = current_location_id()
    active_cases = db.execute(
        f"""
        SELECT c.case_code, c.case_name, c.is_virtual
        FROM cases c
        WHERE c.is_active=1 AND c.location_id=?
        {case_order_sql()}
        """,
        (location_id,),
    ).fetchall()

    if request.method == "POST":
        from_case = (request.form.get("from_case_code") or "").strip()
        to_case = (request.form.get("to_case_code") or "").strip()
        description = (request.form.get("description") or "").strip() or None
        upc_map = parse_upc_lines(request.form.get("upcs", ""))
        picked_map = parse_upc_lines(request.form.get("upcs_picked", ""))
        for k, v in picked_map.items():
            upc_map[k] = upc_map.get(k, 0) + v
        if not from_case or not to_case or not upc_map:
            flash("From case, To case, and at least one UPC are required.", "danger")
            return redirect(url_for("move"))
        if from_case == to_case:
            flash("From and To case can’t be the same.", "danger")
            return redirect(url_for("move"))
        if not ensure_case_exists(from_case) or not ensure_case_exists(to_case):
            flash("From/To case not found.", "danger")
            return redirect(url_for("move"))

        problems = _validate_have_qty(from_case, upc_map, LOCATION_CASE)
        if problems:
            flash("Not enough quantity to move for: " + "; ".join(problems), "danger")
            return redirect(url_for("move"))

        for upc, qty in upc_map.items():
            upsert_product(upc, description, item_type=None)
            ok, _ = remove_qty(from_case, upc, qty, LOCATION_CASE)
            if ok:
                add_qty(to_case, upc, qty, LOCATION_CASE)
                log_history(ACTION_MOVE, upc=upc, qty=qty, from_case_code=from_case, to_case_code=to_case, notes="Moved qty (bulk move page)")
        db.commit()

        flash(f"Moved {sum(upc_map.values())} unit(s) from {from_case} → {to_case}.", "success")
        return redirect(url_for("view_case", case_code=to_case))

    return render_template(
        "move.html",
        active_cases=active_cases,
        user=current_user(),
        new_receipts_code=new_receipts_case_code(location_id),
    )


# ---------------- Standalone Sell/Missing ----------------
@app.route("/sell", methods=["GET", "POST"])
@login_required
def sell():
    db = get_db()
    location_id = current_location_id()
    active_cases = db.execute(
        f"""
        SELECT c.case_code, c.case_name, c.is_virtual
        FROM cases c
        WHERE c.is_active=1 AND c.location_id=?
        {case_order_sql()}
        """,
        (location_id,),
    ).fetchall()

    if request.method == "POST":
        upc = (request.form.get("upc") or "").strip()
        try:
            qty = int((request.form.get("qty") or "1").strip())
        except ValueError:
            flash("Qty must be a whole number greater than 0.", "danger")
            return redirect(url_for("sell"))
        case_code = (request.form.get("case_code") or "").strip() or None

        if not upc or qty <= 0:
            flash("UPC required and qty must be > 0.", "danger")
            return redirect(url_for("sell"))
        if not case_code or not ensure_case_exists(case_code):
            flash("Please choose a valid case.", "danger")
            return redirect(url_for("sell"))

        sold_fields, error = parse_sold_fields(request.form)
        if error:
            flash(error, "danger")
            return redirect(url_for("sell"))

        ok, have = remove_qty(case_code, upc, qty, LOCATION_CASE)
        if not ok:
            flash(f"Not enough qty in case {case_code}. Requested {qty}, have {have}.", "danger")
            return redirect(url_for("sell"))

        db.commit()
        log_history(
            ACTION_SOLD,
            upc=upc,
            qty=qty,
            from_case_code=case_code,
            notes="Sold (standalone)",
            trans_reg=sold_fields["trans_reg"],
            dept_no=sold_fields["dept_no"],
            brief_desc=sold_fields["brief_desc"],
            ticket_price=sold_fields["ticket_price"],
            diamond_test=sold_fields["diamond_test"],
        )
        flash(f"Sold {qty} unit(s) of {upc} from case {case_code}.", "success")
        return redirect(url_for("view_case", case_code=case_code))

    return render_template("sell.html", active_cases=active_cases, user=current_user())


@app.route("/missing", methods=["GET", "POST"])
@login_required
def missing():
    db = get_db()
    location_id = current_location_id()
    active_cases = db.execute(
        f"""
        SELECT c.case_code, c.case_name, c.is_virtual
        FROM cases c
        WHERE c.is_active=1 AND c.location_id=?
        {case_order_sql()}
        """,
        (location_id,),
    ).fetchall()

    if request.method == "POST":
        upc = (request.form.get("upc") or "").strip()
        try:
            qty = int((request.form.get("qty") or "1").strip())
        except ValueError:
            flash("Qty must be a whole number greater than 0.", "danger")
            return redirect(url_for("missing"))
        case_code = (request.form.get("case_code") or "").strip() or None
        notes = (request.form.get("notes") or "").strip() or None

        if not upc or qty <= 0:
            flash("UPC required and qty must be > 0.", "danger")
            return redirect(url_for("missing"))
        if not case_code or not ensure_case_exists(case_code):
            flash("Please choose a valid case.", "danger")
            return redirect(url_for("missing"))

        ok, have = remove_qty(case_code, upc, qty, LOCATION_CASE)
        if not ok:
            flash(f"Not enough qty in case {case_code}. Requested {qty}, have {have}.", "danger")
            return redirect(url_for("missing"))

        db.commit()
        log_history(ACTION_MISSING, upc=upc, qty=qty, from_case_code=case_code, notes=notes or "Missing (standalone)")
        flash(f"Marked MISSING: {qty} unit(s) of {upc} from case {case_code}.", "success")
        return redirect(url_for("view_case", case_code=case_code))

    return render_template("missing.html", active_cases=active_cases, user=current_user())


# ---------------- History ----------------
@app.route("/history")
@login_required
def history():
    db = get_db()
    location_id = current_location_id()
    report_type = (request.args.get("report") or "events").strip().lower()
    if report_type not in ("events", "counts"):
        report_type = "events"

    case_code = (request.args.get("case_code") or "").strip()
    upc = (request.args.get("upc") or "").strip()
    action = (request.args.get("action") or "").strip()
    date = (request.args.get("date") or "").strip()  # YYYY-MM-DD (store-local) for counts

    active_cases = db.execute(
        f"SELECT c.* FROM cases c WHERE c.is_active=1 AND c.location_id=? {case_order_sql()}",
        (location_id,),
    ).fetchall()

    if report_type == "counts":
        sql = "SELECT * FROM case_counts WHERE location_id=?"
        params = [location_id]
        if case_code:
            sql += " AND case_code=?"
            params.append(case_code)
        if date:
            sql += " AND local_date=?"
            params.append(date)
        sql += " ORDER BY id DESC LIMIT 500"
        rows = db.execute(sql, params).fetchall()
        # Current system totals (used to show variance in the report)
        sys_totals_counts = {
            r["case_code"]: case_type_totals(r["case_code"], location_id=location_id)
            for r in rows
        }

        return render_template(
            "history.html",
            report_type=report_type,
            rows=rows,
            sys_totals_counts=sys_totals_counts,
            active_cases=active_cases,
            case_code=case_code,
            upc=upc,
            action=action,
            date=date,
            count_categories=COUNT_CATEGORIES,
            user=current_user(),
        )

    sql = """
        SELECT h.*, p.item_type
        FROM history h
        LEFT JOIN products p ON p.upc = h.upc
        WHERE h.location_id=?
    """
    params = [location_id]
    if case_code:
        sql += " AND (h.from_case_code=? OR h.to_case_code=?)"
        params.extend([case_code, case_code])
    if upc:
        sql += " AND h.upc=?"
        params.append(upc)
    if action:
        sql += " AND h.action LIKE ?"
        params.append(action)

    sql += " ORDER BY id DESC LIMIT 500"
    rows = db.execute(sql, params).fetchall()

    return render_template(
        "history.html",
        report_type=report_type,
        rows=rows,
        active_cases=active_cases,
        case_code=case_code,
        upc=upc,
        action=action,
        date=date,
        count_categories=COUNT_CATEGORIES,
        user=current_user(),
    )



@app.route("/export/inventory.csv")
@login_required
def export_inventory():
    """Export current inventory (by case/upc/qty)."""
    db = get_db()
    location_id = current_location_id()
    rows = db.execute(
        """
        SELECT
          i.case_code,
          c.case_name,
          i.location,
          i.upc,
          COALESCE(p.item_type,'') AS item_type,
          COALESCE(p.description,'') AS description,
          i.qty
        FROM inventory i
        JOIN cases c ON c.case_code = i.case_code AND c.location_id = i.location_id
        LEFT JOIN products p ON p.upc = i.upc
        WHERE c.is_active = 1 AND i.qty > 0 AND c.location_id = ?
        ORDER BY i.case_code, i.location, i.upc
        """,
        (location_id,),
    ).fetchall()

    import io, csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["case_code","case_name","location","upc","item_type","description","qty"])
    for r in rows:
        w.writerow([r["case_code"], r["case_name"], r["location"], r["upc"], r["item_type"], r["description"], int(r["qty"])])

    data = buf.getvalue().encode("utf-8")
    filename = f"inventory_{now_local_dt().strftime('%m-%d-%Y_%H%M')}.csv"
    resp = make_response(data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


@app.route("/export/case/<case_code>.csv")
@login_required
def export_case(case_code):
    """Export inventory for a single case."""
    db = get_db()
    location_id = current_location_id()
    c = db.execute(
        "SELECT case_code, case_name FROM cases WHERE case_code=? AND is_active=1 AND location_id=?",
        (case_code, location_id),
    ).fetchone()
    if not c:
        abort(404)

    rows = db.execute(
        """
        SELECT i.upc,
               i.location,
               COALESCE(p.item_type,'') AS item_type,
               COALESCE(p.description,'') AS description,
               i.qty
        FROM inventory i
        LEFT JOIN products p ON p.upc = i.upc
        WHERE i.case_code = ? AND i.location_id = ? AND i.qty > 0
        ORDER BY i.location, i.upc
        """,
        (case_code, location_id),
    ).fetchall()

    import io, csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["case_code","case_name","location","upc","item_type","description","qty"])
    for r in rows:
        w.writerow([c["case_code"], c["case_name"], r["location"], r["upc"], r["item_type"], r["description"], int(r["qty"])])

    data = buf.getvalue().encode("utf-8")
    filename = f"case_{case_code}_{now_local_dt().strftime('%m-%d-%Y_%H%M')}.csv"
    resp = make_response(data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp
@app.route("/export/history.csv")
@login_required
def export_history():
    db = get_db()
    location_id = current_location_id()
    report_type = (request.args.get("report") or "events").strip().lower()
    if report_type not in ("events", "counts"):
        report_type = "events"

    case_code = (request.args.get("case_code") or "").strip()
    upc = (request.args.get("upc") or "").strip()
    action = (request.args.get("action") or "").strip()
    date = (request.args.get("date") or "").strip()

    import io, csv

    if report_type == "counts":
        sql = "SELECT * FROM case_counts WHERE location_id=?"
        params = [location_id]
        if case_code:
            sql += " AND case_code=?"
            params.append(case_code)
        if date:
            sql += " AND local_date=?"
            params.append(date)
        sql += " ORDER BY id DESC LIMIT 5000"
        rows = db.execute(sql, params).fetchall()

        output = io.StringIO()
        w = csv.writer(output)
        w.writerow([
            "ts_utc",
            "local_date",
            "case_code",
            "username",
            "bracelets",
            "rings",
            "earrings",
            "necklaces",
            "other",
            "reserve_bracelets",
            "reserve_rings",
            "reserve_earrings",
            "reserve_necklaces",
            "reserve_other",
            "total",
            "notes",
        ])
        for r in rows:
            w.writerow([
                r["ts_utc"], r["local_date"], r["case_code"], r["username"] or "",
                r["bracelets"], r["rings"], r["earrings"], r["necklaces"], r["other"],
                r["reserve_bracelets"], r["reserve_rings"], r["reserve_earrings"],
                r["reserve_necklaces"], r["reserve_other"],
                r["total"],
                r["notes"] or ""
            ])
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=counts_history.csv"},
        )

    sql = "SELECT h.*, p.item_type FROM history h LEFT JOIN products p ON p.upc = h.upc WHERE h.location_id=?"
    params = [location_id]
    if case_code:
        sql += " AND (h.from_case_code=? OR h.to_case_code=?)"
        params.extend([case_code, case_code])
    if upc:
        sql += " AND h.upc=?"
        params.append(upc)
    if action:
        sql += " AND h.action LIKE ?"
        params.append(action)
    sql += " ORDER BY h.id DESC LIMIT 5000"
    rows = db.execute(sql, params).fetchall()

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["ts","username","action","upc","item_type","qty","from_case_code","to_case_code","notes","trans_reg","dept_no","brief_desc","ticket_price","diamond_test"])
    for r in rows:
        w.writerow([
            r["ts"], r["username"] or "", r["action"], r["upc"] or "",
            r["item_type"] or "", r["qty"] or "",
            r["from_case_code"] or "", r["to_case_code"] or "", r["notes"] or "",
            r["trans_reg"] or "", r["dept_no"] or "", r["brief_desc"] or "",
            r["ticket_price"] if (r["ticket_price"] is not None) else "",
            r["diamond_test"] or ""
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=events_history.csv"},
    )


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def users():
    db = get_db()
    locations = db.execute(
        "SELECT id, name FROM locations WHERE is_active=1 ORDER BY name"
    ).fetchall()

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        role = (request.form.get("role") or ROLE_STAFF).strip()
        location_id_raw = (request.form.get("location_id") or "").strip() or None

        if not username:
            flash("Username is required.", "danger")
            return redirect(url_for("users"))
        if len(password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for("users"))
        if role not in (ROLE_ADMIN, ROLE_STAFF):
            role = ROLE_STAFF

        location_id = None
        if location_id_raw:
            try:
                location_id = int(location_id_raw)
            except ValueError:
                flash("Choose a valid location.", "danger")
                return redirect(url_for("users"))
        if role == ROLE_STAFF and not location_id:
            flash("Staff users must have a location.", "danger")
            return redirect(url_for("users"))
        if location_id:
            exists = db.execute(
                "SELECT id FROM locations WHERE id=? AND is_active=1",
                (location_id,),
            ).fetchone()
            if not exists:
                flash("Choose a valid location.", "danger")
                return redirect(url_for("users"))

        existing = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
        if existing:
            flash("That username already exists.", "danger")
            return redirect(url_for("users"))

        db.execute(
            """
            INSERT INTO users (username, password_hash, role, is_active, location_id, created_at)
            VALUES (?,?,?,?,?,?)
            """,
            (username, generate_password_hash(password), role, 1, location_id, utc_now()),
        )
        db.commit()
        log_history(ACTION_USER_CREATE, notes=f"Created user {username} ({role})")
        flash("User created.", "success")
        return redirect(url_for("users"))

    rows = db.execute(
        """
        SELECT u.id, u.username, u.role, u.is_active, u.created_at, l.name AS location_name
        FROM users u
        LEFT JOIN locations l ON l.id = u.location_id
        ORDER BY u.id
        """
    ).fetchall()
    return render_template(
        "users.html",
        rows=rows,
        locations=locations,
        user=current_user(),
    )


@app.route("/admin/users/<int:user_id>/disable", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def disable_user(user_id: int):
    db = get_db()
    me = current_user()
    if me and me.id == user_id:
        flash("You can’t disable yourself. (Nice try though.)", "danger")
        return redirect(url_for("users"))

    db.execute("UPDATE users SET is_active=0 WHERE id=?", (user_id,))
    db.commit()
    log_history(ACTION_USER_DISABLE, notes=f"Disabled user_id={user_id}")
    flash("User disabled.", "success")
    return redirect(url_for("users"))


@app.route("/reports/daily", methods=["GET"])
@login_required
def daily_activity_reports():
    # Choose date and case, then download report
    db = get_db()
    location_id = current_location_id()
    cases = db.execute(
        "SELECT case_code, case_name FROM cases WHERE is_active = 1 AND location_id = ? ORDER BY case_code",
        (location_id,),
    ).fetchall()
    # default date in store local
    default_date = local_date_str()
    date_q = request.args.get("date") or default_date
    # Accept MM/DD/YYYY too
    if re.match(r"^\d{2}/\d{2}/\d{4}$", date_q):
        mm, dd, yyyy = date_q.split("/")
        date_q = f"{yyyy}-{mm}-{dd}"
    return render_template("daily_reports.html", cases=cases, date=date_q, user=current_user())

@app.route("/reports/daily/<case_code>.xlsx")
@login_required
def download_daily_activity_report(case_code):
    date_q = request.args.get("date") or local_date_str()
    if re.match(r"^\d{2}/\d{2}/\d{4}$", date_q):
        mm, dd, yyyy = date_q.split("/")
        date_q = f"{yyyy}-{mm}-{dd}"

    # Ensure case exists
    db = get_db()
    location_id = current_location_id()
    c = db.execute(
        "SELECT case_code, case_name FROM cases WHERE case_code = ? AND is_active = 1 AND location_id = ?",
        (case_code, location_id),
    ).fetchone()
    if not c:
        abort(404)

    wb = build_daily_activity_workbook(case_code, date_q, location_id)

    import io
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    safe_date = date_q.replace("-", "")
    filename = f"Daily_Activity_{case_code}_{safe_date}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/reports/daily-counts/<case_code>.xlsx")
@login_required
def download_daily_count_report(case_code):
    date_q = request.args.get("date") or local_date_str()
    if re.match(r"^\d{2}/\d{2}/\d{4}$", date_q):
        mm, dd, yyyy = date_q.split("/")
        date_q = f"{yyyy}-{mm}-{dd}"

    db = get_db()
    location_id = current_location_id()
    c = db.execute(
        "SELECT case_code, case_name FROM cases WHERE case_code = ? AND is_active = 1 AND location_id = ?",
        (case_code, location_id),
    ).fetchone()
    if not c:
        abort(404)

    wb = build_daily_count_workbook(case_code, date_q, location_id)

    import io
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    safe_date = date_q.replace("-", "")
    filename = f"Daily_Count_{case_code}_{safe_date}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if __name__ == "__main__":
    init_db()
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "1").strip() == "1"
    use_https = os.getenv("USE_HTTPS", "1").strip() == "1"
    cert_file = os.getenv("SSL_CERT_FILE")
    key_file = os.getenv("SSL_KEY_FILE")
    ssl_context = None
    if use_https:
        if cert_file and key_file:
            ssl_context = (cert_file, key_file)
        else:
            ssl_context = "adhoc"
        app.config.update(
            PREFERRED_URL_SCHEME="https",
            SESSION_COOKIE_SECURE=True,
        )
    app.run(host=host, port=port, debug=debug, ssl_context=ssl_context)
